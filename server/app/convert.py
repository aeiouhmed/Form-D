"""
Utilities for transforming source spreadsheets into the K1 import format.

This module contains functions for reading, cleaning, and transforming data
from an uploaded spreadsheet to match the required format for a K1 import.
It handles various data cleaning tasks, including normalizing headers,
sanitizing cell values, and mapping HS codes to their corresponding units.
"""

from __future__ import annotations

from io import BytesIO
import logging
import os
import re
from pathlib import Path
from typing import Iterable, List
import json  # Added json import

import pandas as pd
from openpyxl import load_workbook
import xlrd
from xlutils.copy import copy as xl_copy
import xlwt

# Paths
TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "templates" / "K1 Import Template.xls"
# CORRECTED: Point to the JSON file
HS_MAPPING_PATH = Path(__file__).resolve().parent.parent / "templates" / "HSCODE.json"

ALWAYS_BLANK_COLUMN_NAMES = [
    "ExciseDutyMethod",
    "ExciseDutyRateExemptedPercentage",
    "ExciseDutyRateExemptedSpecific",
    "VehicleType",
    "VehicleModel",
    "Brand",
    "Engine",
    "Chassis",
    "CC",
    "Year",
]

# Define lists of possible header names for critical columns.
HS_CANDIDATES = ["Hs Code", "HS Code", "Hscode", "HSCode", "HS-Code"]
HS_MAPPING_HEADER_VARIANTS = [
    "Hs Code",
    "HS Code",
    "Hscode",
    "HSCode",
    "HsCode",
    "H S Code",
    "Hs code",
    "HS code",
]
UNIT_HEADER_VARIANTS = ["Unit", "Units", "UOM", "UNTnit"]
NET_WEIGHT_CANDIDATES = [
    "Net Weight(Kg)",
    "Net Weight (Kg)",
    "Net Weight",
    "Weight (Kg)",
    "Weight",
]
AMOUNT_CANDIDATES = ["Amount(USD)", "Amount (USD)", "Amount USD", "Amount"]
PARTS_CANDIDATES = ["Parts Name", "Description", "Item Description"]
QUANTITY_CANDIDATES = ["Quantity", "Qty", "QTY"]
FORM_FLAG_CANDIDATES = ["Form Flag", "FormFlag", "Form_Flag", "Form flag"]

# Pre-compile a regular expression to find and remove bad control characters.
_CTRL_BAD = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")


def _normalize_header(value: object) -> str:
    """Normalize a header by removing whitespace, hyphens, and converting to lowercase."""
    return re.sub(r"[\s_\-]+", "", str(value or "").strip().lower())


def _digits_only(value: object) -> str:
    """Extract only the digits from a string."""
    return re.sub(r"\D", "", str(value or ""))


def _hs_out_code(value: object) -> str:
    """Format an HS code by extracting digits and appending '00'."""
    digits = _digits_only(value)
    return f"{digits}00"


def _sanitize_cell(value: object) -> object:
    """
    Sanitize a cell's value by removing control characters and truncating if necessary.
    """
    if value is None or pd.isna(value):
        return ""
    if hasattr(value, "item"):
        value = value.item()
    if isinstance(value, (int, float)):
        return value
    text = str(value)
    cleaned = _CTRL_BAD.sub(" ", text)
    if len(cleaned) > 32767:
        cleaned = cleaned[:32767]
    return cleaned


def _load_template_columns_xlsx(path: Path) -> list[str]:
    """Load the header row from an XLSX template file."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers: list[str] = []
    first_row = next(ws.iter_rows(min_row=1, max_row=1), ())
    for cell in first_row:
        value = cell.value
        if isinstance(value, str):
            value = value.strip().lstrip("'")
        headers.append("" if value is None else str(value))
    wb.close()
    return headers


def _load_template_columns_xls(path: Path) -> list[str]:
    """Load the header row from the 'JobCargo' sheet in an XLS template file."""
    book = xlrd.open_workbook(path)
    try:
        sheet = book.sheet_by_name("JobCargo")
    except xlrd.biffh.XLRDError:
        raise ValueError("Sheet 'JobCargo' not found in the template.")

    headers: list[str] = []
    for col_idx in range(sheet.ncols):
        value = sheet.cell_value(0, col_idx)
        if isinstance(value, str):
            value = value.strip().lstrip("'")
        headers.append("" if value is None else str(value))
    return headers


# CORRECTED: New JSON loading function
def _load_hs_mapping() -> dict[str, str]:
    """
    Load the HS code to unit mapping from the JSON file.
    """
    if not HS_MAPPING_PATH.exists():
        # Fallback logic or error if file is missing
        logging.warning(f"HSCODE.json not found at {HS_MAPPING_PATH}")
        return {}
        
    try:
        with open(HS_MAPPING_PATH, "r", encoding="utf-8") as f:
            mapping = json.load(f)
            
        logging.info("Loaded HS mapping with %s entries.", len(mapping))
        return mapping
    except Exception as exc:
        raise RuntimeError(f"Failed to load HSCODE.json: {exc}") from exc


HS_CODE_TO_UNIT = _load_hs_mapping()
DEBUG_HSLOOKUP = os.getenv("DEBUG_HSLOOKUP") == "1"
_LOGGED_MISSING_CODES: set[tuple[str, str, tuple[str, ...]]] = set()


def normalize_for_match(value: object) -> str:
    """Alias for _normalize_header for clarity in matching contexts."""
    return _normalize_header(value)


ALWAYS_BLANK_NORMALIZED = {
    normalize_for_match(column_name) for column_name in ALWAYS_BLANK_COLUMN_NAMES
}
ALWAYS_BLANK_COLLAPSED = {
    re.sub(r"[^a-z0-9]", "", value) for value in ALWAYS_BLANK_NORMALIZED
}


def _to_xls_bytes_with_template(final_df: pd.DataFrame, template_path: Path) -> bytes:
    """Convert a DataFrame to XLS bytes using a template."""
    if not template_path.exists():
        raise FileNotFoundError(f"Template file not found at {template_path}")

    book_reader = xlrd.open_workbook(template_path, formatting_info=True)
    book_writer = xl_copy(book_reader)

    sheet_writer = None
    try:
        sheet_names = book_reader.sheet_names()
        if "JobCargo" not in sheet_names:
            raise ValueError("Sheet 'JobCargo' not found in the template.")
        sheet_index = sheet_names.index("JobCargo")
        sheet_writer = book_writer.get_sheet(sheet_index)
    except IndexError:
        raise ValueError("Sheet 'JobCargo' not found in the template.")

    data_to_write = final_df.values.tolist()

    for r_idx, row_data in enumerate(data_to_write, start=1):
        for c_idx, value in enumerate(row_data):
            sanitized_value = _sanitize_cell(value)
            sheet_writer.write(r_idx, c_idx, sanitized_value)

    buffer = BytesIO()
    book_writer.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def convert_to_k1(
    uploaded_bytes: bytes,
    country: str = "ID",
    template_path: str | Path | None = None,
) -> bytes:
    """
    Convert uploaded spreadsheet bytes into K1 import XLSX bytes.
    """
    source_df = _load_source_dataframe(uploaded_bytes)
    source_df = _normalize_columns(source_df)

    form_flag_col = _first_matching_col(source_df.columns, FORM_FLAG_CANDIDATES)
    if form_flag_col is None:
        raise ValueError("Missing required column: 'Form Flag'")

    total_rows = len(source_df)
    mask = source_df[form_flag_col].apply(_normalize_flag).eq("form d")
    kept_rows = int(mask.sum())
    logging.info("Form Flag filter: kept %s of %s rows", kept_rows, total_rows)
    if kept_rows == 0:
        raise ValueError("No rows with 'Form D' in 'Form Flag'.")
    source_df = source_df.loc[mask].reset_index(drop=True)

    template_file = _resolve_template_path(template_path)
    logging.info("Using template at: %s", template_file)
    template_columns = _load_template(template_file)

    log_config(list(source_df.columns), template_columns)

    output = pd.DataFrame(index=source_df.index)

    hs_series = _get_series(source_df, HS_CANDIDATES, default="").fillna("")
    hs_clean = hs_series.apply(_hs_out_code)

    units: list[str] = []
    prefix_lists: list[list[str]] = []
    for raw_hs, hs_out in zip(hs_series, hs_clean):
        hs_out_str = str(hs_out or "")
        max_len = min(len(hs_out_str), 10)
        prefixes: list[str] = []
        for length in range(max_len, 4, -1):
            key = hs_out_str[:length]
            if not key or (prefixes and key == prefixes[-1]):
                continue
            prefixes.append(key)
        if not prefixes:
            prefixes.append(hs_out_str)
        unit = "N/A"
        matched_key = None
        for candidate in prefixes:
            mapping_value = HS_CODE_TO_UNIT.get(candidate)
            if mapping_value is None:
                continue
            normalized_unit = str(mapping_value or "").strip().upper()
            if normalized_unit in {"", "NA", "NAN", "NULL"}:
                continue
            unit = normalized_unit
            matched_key = candidate
            break
        if DEBUG_HSLOOKUP and matched_key and matched_key != hs_out_str:
            logging.info(
                "HS matched on prefix: raw='%s' chosen='%s' unit='%s'",
                raw_hs,
                matched_key,
                unit,
            )
        units.append(unit)
        prefix_lists.append(prefixes)
    unit_series = pd.Series(units, index=source_df.index, dtype="object")

    quantity_series = pd.to_numeric(
        _get_series(source_df, QUANTITY_CANDIDATES, default=0), errors="coerce"
    ).fillna(0)
    net_weight_series = pd.to_numeric(
        _get_series(source_df, NET_WEIGHT_CANDIDATES, default=0),
        errors="coerce",
    ).fillna(0)
    amount_series = pd.to_numeric(
        _get_series(source_df, AMOUNT_CANDIDATES, default=0),
        errors="coerce",
    ).fillna(0)
    parts_name_series = _get_series(
        source_df, PARTS_CANDIDATES, default=""
    ).fillna("")
    parts_name_series = parts_name_series.astype(str)

    unit_upper = unit_series.str.upper()

    statistical_qty = pd.Series([""] * len(source_df), index=source_df.index, dtype="object")
    declared_qty = pd.Series([""] * len(source_df), index=source_df.index, dtype="object")

    kgm_mask = unit_upper == "KGM"
    if kgm_mask.any():
        net_values = net_weight_series.loc[kgm_mask].astype(float)
        statistical_qty.loc[kgm_mask] = net_values
        declared_qty.loc[kgm_mask] = net_values

    unt_mask = unit_upper == "UNT"
    if unt_mask.any():
        qty_values = quantity_series.loc[unt_mask].astype(float)
        statistical_qty.loc[unt_mask] = qty_values
        declared_qty.loc[unt_mask] = qty_values

    if DEBUG_HSLOOKUP and (unit_upper == "N/A").any():
        for position, (_, value) in enumerate(unit_series.items()):
            if str(value).upper() != "N/A":
                continue
            raw_code = hs_series.iloc[position]
            hs_out_value = hs_clean.iloc[position]
            prefixes = tuple(prefix_lists[position])
            key = (str(raw_code), str(hs_out_value), prefixes)
            if key not in _LOGGED_MISSING_CODES:
                logging.warning(
                    "HS code not in mapping: raw='%s' -> hs_out='%s' candidates=%s",
                    raw_code,
                    hs_out_value,
                    list(prefixes),
                )
                _LOGGED_MISSING_CODES.add(key)

    country_value = (country or "ID").strip().upper() or "ID"

    output["CountryOfOrigin"] = country_value
    output["HSCode"] = hs_clean
    output["StatisticalUOM"] = unit_series
    output["DeclaredUOM"] = unit_series
    output["StatisticalQty"] = statistical_qty
    output["DeclaredQty"] = declared_qty
    output["ItemAmount"] = amount_series
    output["ItemDescription"] = parts_name_series
    output["ItemDescription2"] = quantity_series
    output["ItemDescription3"] = ""

    output["ImportDutyMethod"] = "Exemption"
    output["ImportDutyRateExemptedPercentage"] = 100
    output["ImportDutyRateExemptedSpecific"] = ""

    output["SSTMethod"] = "Exemption"
    output["SSTRateExemptedPercentage"] = 100
    output["SSTRateExemptedSpecific"] = ""

    output["ExciseDutyMethod"] = ""
    output["ExciseDutyRateExemptedPercentage"] = ""
    output["ExciseDutyRateExemptedSpecific"] = ""

    output["VehicleType"] = ""
    output["VehicleModel"] = ""
    output["Brand"] = ""
    output["Engine"] = ""
    output["Chassis"] = ""
    output["CC"] = ""
    output["Year"] = ""

    if "StatisticalUOM" not in output.columns:
        output["StatisticalUOM"] = pd.Series([""] * len(output), index=output.index, dtype="object")
    if "DeclaredUOM" not in output.columns:
        output["DeclaredUOM"] = output["StatisticalUOM"].copy()

    assert (output["StatisticalUOM"] == output["DeclaredUOM"]).all(), (
        "DeclaredUOM must match StatisticalUOM"
    )

    normalized_output_map = {
        normalize_for_match(column): output[column] for column in output.columns
    }
    collapsed_output_map = {
        re.sub(r"[^a-z0-9]", "", key): value for key, value in normalized_output_map.items()
    }

    method_occurrence = 0
    final_series: list[pd.Series] = []
    for template_column in template_columns:
        normalized_template_column = _normalize_header(template_column)
        normalized_template_key = normalize_for_match(template_column)
        collapsed_template_key = re.sub(r"[^a-z0-9]", "", normalized_template_key)

        if normalize_for_match(normalized_template_column) == "method":
            method_occurrence += 1
            if method_occurrence in (1, 2):
                fill_value = "E"
            elif method_occurrence == 3:
                fill_value = ""
            else:
                fill_value = ""
            series = pd.Series(
                [fill_value] * len(output),
                index=output.index,
                dtype="object",
            )
        elif (
            normalized_template_key in ALWAYS_BLANK_NORMALIZED
            or collapsed_template_key in ALWAYS_BLANK_COLLAPSED
        ):
            series = pd.Series([""] * len(output), index=output.index, dtype="object")
        elif normalized_template_key in normalized_output_map:
            series = normalized_output_map[normalized_template_key]
        elif collapsed_template_key in collapsed_output_map:
            series = collapsed_output_map[collapsed_template_key]
        else:
            series = pd.Series([""] * len(output), index=output.index, dtype="object")

        final_series.append(series.rename(template_column))

    final_df = (
        pd.concat(final_series, axis=1) if final_series else pd.DataFrame(index=output.index)
    )

    _maybe_log_debug_samples(final_df)

    template_file = _resolve_template_path(template_path)
    return _to_xls_bytes_with_template(final_df, template_file)


def _load_source_dataframe(uploaded_bytes: bytes) -> pd.DataFrame:
    """
    Load a spreadsheet from bytes into a pandas DataFrame.
    """
    buffer = BytesIO(uploaded_bytes)
    buffer.seek(0)
    head = buffer.read(4)
    buffer.seek(0)
    is_xlsx_like = head.startswith(b"PK")
    if is_xlsx_like:
        return pd.read_excel(buffer, engine="openpyxl")
    return _df_from_xls_bytes(uploaded_bytes)


def _load_template(template_path: Path) -> List[str]:
    """Load the template columns from either an XLS or XLSX file."""
    ext = template_path.suffix.lower()
    if ext in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _load_template_columns_xlsx(template_path)
    if ext == ".xls":
        return _load_template_columns_xls(template_path)
    raise ValueError(f"Unsupported template format: {template_path}")


def _resolve_template_path(custom_path: str | Path | None) -> Path:
    """Resolve the path to the template file, using a default if none is provided."""
    if custom_path is None:
        path = TEMPLATE_PATH
    else:
        path = Path(custom_path)
        if not path.is_absolute():
            path = path.resolve()
    if not path.exists():
        raise ValueError(f"Template not found at {path.resolve()}")
    return path.resolve()


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize all column headers in a DataFrame."""
    df = df.copy()
    normalized = [normalize_for_match(col) for col in df.columns]
    df.columns = normalized
    return df


def _get_series(
    df: pd.DataFrame,
    candidates: Iterable[str],
    default: object,
) -> pd.Series:
    """
    Get a column's data as a pandas Series by searching for candidate column names.
    """
    column_name = _first_matching_col(df.columns, candidates)
    if column_name is None:
        return pd.Series([default] * len(df), index=df.index)
    return df[column_name]


def log_config(source_columns: list[str], template_columns: list[str]) -> None:
    """Log column configuration for debugging purposes."""
    logging.info("Source columns (normalized): %s", source_columns)
    logging.info("Template columns (original): %s", template_columns)


def _first_matching_col(columns: Iterable[str], candidates: Iterable[str]) -> str | None:
    """
    Find the first column name that matches any of the candidates.
    """
    columns_list = list(columns)
    normalized_columns = [normalize_for_match(col) for col in columns_list]
    candidate_keys = [normalize_for_match(candidate) for candidate in candidates]
    # Exact match on normalized names.
    for candidate_key in candidate_keys:
        for idx, column_key in enumerate(normalized_columns):
            if column_key == candidate_key:
                return columns_list[idx]
    # Substring match on normalized names.
    for candidate_key in candidate_keys:
        if not candidate_key:
            continue
        for idx, column_key in enumerate(normalized_columns):
            if candidate_key in column_key:
                return columns_list[idx]
    # Match on "collapsed" names (alphanumeric only).
    collapsed_columns = [re.sub(r"[^a-z0-9]", "", column_key) for column_key in normalized_columns]
    collapsed_candidates = [re.sub(r"[^a-z0-9]", "", candidate_key) for candidate_key in candidate_keys]
    for candidate_key in collapsed_candidates:
        if not candidate_key:
            continue
        for idx, column_key in enumerate(collapsed_columns):
            if column_key == candidate_key or candidate_key in column_key:
                return columns_list[idx]
    return None


def _normalize_flag(value: object) -> str:
    """Normalize the 'Form Flag' value for consistent matching."""
    if value is None or pd.isna(value):
        return ""
    text = str(value).lower()
    text = re.sub(r"[-_]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _sanitize_hs(value: object) -> str:
    """Sanitize an HS code value."""
    return _hs_out_code(value)


def _maybe_log_debug_samples(df: pd.DataFrame) -> None:
    """Log a few sample rows of the final DataFrame for debugging if enabled."""
    if os.environ.get("DEBUG_MAPPINGS") == "1":
        records = df.head(3).to_dict(orient="records")
        logging.info("Sample mapped rows: %s", records)


def _df_from_xls_bytes(xls_bytes: bytes) -> pd.DataFrame:
    """Create a pandas DataFrame from the bytes of an XLS file."""
    import xlrd

    book = xlrd.open_workbook(file_contents=xls_bytes)
    sheet = book.sheet_by_index(0)
    headers = [str(sheet.cell_value(0, c)).lstrip("'").strip() for c in range(sheet.ncols)]
    rows: list[list[object]] = []
    for r in range(1, sheet.nrows):
        row_values = [sheet.cell_value(r, c) for c in range(sheet.ncols)]
        rows.append(row_values)
    if rows:
        return pd.DataFrame(rows, columns=headers)
    return pd.DataFrame(columns=headers)